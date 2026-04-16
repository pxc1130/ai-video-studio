#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Lightweight SQLite-based batch job queue for ai-video-studio.
Suitable for 1-3 concurrent users, processing up to ~150 jobs/day.
"""

from __future__ import annotations

import csv
import io
import json
import os
import sqlite3
import threading
import time
import uuid
import zipfile
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

# DB path relative to this file -> project_root/backend_data/batch_queue.db
DB_PATH = Path(__file__).resolve().parent.parent.parent / "backend_data" / "batch_queue.db"
DB_PATH.parent.mkdir(parents=True, exist_ok=True)


def _get_conn() -> sqlite3.Connection:
    conn = sqlite3.connect(str(DB_PATH), check_same_thread=False)
    conn.row_factory = sqlite3.Row
    return conn


def init_db() -> None:
    conn = _get_conn()
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS batches (
            id TEXT PRIMARY KEY,
            status TEXT NOT NULL DEFAULT 'pending',
            total_items INTEGER NOT NULL DEFAULT 0,
            completed_items INTEGER NOT NULL DEFAULT 0,
            failed_items INTEGER NOT NULL DEFAULT 0,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            eta_seconds INTEGER,
            zip_path TEXT,
            error_message TEXT
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS batch_items (
            id TEXT PRIMARY KEY,
            batch_id TEXT NOT NULL,
            idx INTEGER NOT NULL,
            status TEXT NOT NULL DEFAULT 'pending',
            product_id TEXT NOT NULL,
            product_name TEXT NOT NULL,
            price TEXT,
            category TEXT,
            script_template TEXT,
            image_names TEXT,
            product_dir TEXT,
            run_id TEXT,
            plan_path TEXT,
            deliverables_dir TEXT,
            error_message TEXT,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            FOREIGN KEY (batch_id) REFERENCES batches(id) ON DELETE CASCADE
        )
        """
    )
    conn.execute(
        "CREATE INDEX IF NOT EXISTS idx_batch_items_batch_id ON batch_items(batch_id)"
    )
    conn.commit()
    conn.close()


@dataclass
class BatchItemInput:
    product_id: str
    product_name: str
    price: str = ""
    category: str = "default"
    script_template: str = ""
    image_names: list[str] | None = None


class BatchQueue:
    def __init__(self) -> None:
        self._lock = threading.Lock()
        init_db()

    def create_batch(self, items: list[BatchItemInput]) -> str:
        batch_id = uuid.uuid4().hex[:12]
        now = datetime.now().isoformat(timespec="seconds")
        conn = _get_conn()
        with conn:
            conn.execute(
                "INSERT INTO batches (id, status, total_items, created_at, updated_at) VALUES (?, ?, ?, ?, ?)",
                (batch_id, "pending", len(items), now, now),
            )
            for idx, it in enumerate(items):
                conn.execute(
                    """
                    INSERT INTO batch_items (
                        id, batch_id, idx, product_id, product_name, price, category,
                        script_template, image_names, product_dir, run_id, deliverables_dir,
                        created_at, updated_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        f"{batch_id}_{idx}",
                        batch_id,
                        idx,
                        it.product_id,
                        it.product_name,
                        it.price,
                        it.category,
                        it.script_template,
                        json.dumps(it.image_names or [], ensure_ascii=False),
                        "",
                        "",
                        "",
                        now,
                        now,
                    ),
                )
        conn.close()
        return batch_id

    def list_batches(self, limit: int = 50) -> list[dict[str, Any]]:
        conn = _get_conn()
        rows = conn.execute(
            "SELECT * FROM batches ORDER BY created_at DESC LIMIT ?",
            (limit,),
        ).fetchall()
        conn.close()
        return [dict(r) for r in rows]

    def get_batch(self, batch_id: str) -> dict[str, Any] | None:
        conn = _get_conn()
        row = conn.execute("SELECT * FROM batches WHERE id = ?", (batch_id,)).fetchone()
        conn.close()
        return dict(row) if row else None

    def get_batch_items(self, batch_id: str) -> list[dict[str, Any]]:
        conn = _get_conn()
        rows = conn.execute(
            "SELECT * FROM batch_items WHERE batch_id = ? ORDER BY idx",
            (batch_id,),
        ).fetchall()
        conn.close()
        return [dict(r) for r in rows]

    def claim_next_pending_batch(self) -> str | None:
        conn = _get_conn()
        row = conn.execute(
            "SELECT id FROM batches WHERE status = 'pending' ORDER BY created_at LIMIT 1"
        ).fetchone()
        if not row:
            conn.close()
            return None
        batch_id = row["id"]
        now = datetime.now().isoformat(timespec="seconds")
        with conn:
            conn.execute(
                "UPDATE batches SET status = 'running', updated_at = ? WHERE id = ?",
                (now, batch_id),
            )
        conn.close()
        return batch_id

    def update_batch_progress(
        self,
        batch_id: str,
        completed: int | None = None,
        failed: int | None = None,
        status: str | None = None,
        zip_path: str | None = None,
        eta_seconds: int | None = None,
        error_message: str | None = None,
    ) -> None:
        now = datetime.now().isoformat(timespec="seconds")
        conn = _get_conn()
        fields: list[str] = ["updated_at = ?"]
        vals: list[Any] = [now]
        if completed is not None:
            fields.append("completed_items = ?")
            vals.append(completed)
        if failed is not None:
            fields.append("failed_items = ?")
            vals.append(failed)
        if status is not None:
            fields.append("status = ?")
            vals.append(status)
        if zip_path is not None:
            fields.append("zip_path = ?")
            vals.append(zip_path)
        if eta_seconds is not None:
            fields.append("eta_seconds = ?")
            vals.append(eta_seconds)
        if error_message is not None:
            fields.append("error_message = ?")
            vals.append(error_message)
        vals.append(batch_id)
        sql = f"UPDATE batches SET {', '.join(fields)} WHERE id = ?"
        with conn:
            conn.execute(sql, vals)
        conn.close()

    def update_item(
        self,
        item_id: str,
        status: str,
        product_dir: str | None = None,
        run_id: str | None = None,
        plan_path: str | None = None,
        deliverables_dir: str | None = None,
        error_message: str | None = None,
    ) -> None:
        now = datetime.now().isoformat(timespec="seconds")
        conn = _get_conn()
        fields = ["status = ?", "updated_at = ?"]
        vals: list[Any] = [status, now]
        if product_dir is not None:
            fields.append("product_dir = ?")
            vals.append(product_dir)
        if run_id is not None:
            fields.append("run_id = ?")
            vals.append(run_id)
        if plan_path is not None:
            fields.append("plan_path = ?")
            vals.append(plan_path)
        if deliverables_dir is not None:
            fields.append("deliverables_dir = ?")
            vals.append(deliverables_dir)
        if error_message is not None:
            fields.append("error_message = ?")
            vals.append(error_message)
        vals.append(item_id)
        sql = f"UPDATE batch_items SET {', '.join(fields)} WHERE id = ?"
        with conn:
            conn.execute(sql, vals)
        conn.close()

    def get_next_pending_item(self, batch_id: str) -> dict[str, Any] | None:
        conn = _get_conn()
        row = conn.execute(
            "SELECT * FROM batch_items WHERE batch_id = ? AND status = 'pending' ORDER BY idx LIMIT 1",
            (batch_id,),
        ).fetchone()
        conn.close()
        return dict(row) if row else None


# ---------------------------------------------------------------------------
# CSV parser with clear rules
# ---------------------------------------------------------------------------

def _rows_from_spreadsheet(file_bytes: bytes, filename: str) -> list[dict[str, str]]:
    ext = (Path(filename).suffix or "").lower()
    if ext == ".csv":
        text = file_bytes.decode("utf-8-sig", errors="ignore")
        reader = csv.DictReader(io.StringIO(text))
        if not reader.fieldnames:
            raise ValueError("CSV file is empty or has no headers")
        return [{k: (v or "").strip() for k, v in row.items()} for row in reader]

    if ext in (".xlsx", ".xls"):
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        ws = wb.active
        if ws is None:
            raise ValueError("Excel file has no active sheet")
        headers = [str(cell.value or "").strip().lower() for cell in ws[1]]
        rows: list[dict[str, str]] = []
        for raw in ws.iter_rows(min_row=2, values_only=True):
            row = {headers[i]: (str(v or "")).strip() for i, v in enumerate(raw)}
            rows.append(row)
        return rows

    raise ValueError(f"Unsupported file format: {ext}. Please upload .csv or .xlsx/.xls")


def parse_spreadsheet_and_match_images(
    file_bytes: bytes,
    filename: str,
    available_images: list[str],
) -> list[BatchItemInput]:
    """
    Parse a CSV / Excel file and match images according to clear rules:

    Rule 1 (Explicit): If the file contains columns image_1, image_2, image_3...
                       use the filenames written there.
    Rule 2 (Prefix fallback): If no image_xxx columns exist, match files in the
                              same folder whose name starts with product_id.
    Rule 3 (Global pool fallback): If no prefix matches, use all images as a pool
                                   (only if every row has 0 matches, to avoid confusion).
    """
    raw_rows = _rows_from_spreadsheet(file_bytes, filename)
    if not raw_rows:
        raise ValueError("Spreadsheet is empty or has no data rows")

    headers = [h.strip().lower() for h in raw_rows[0].keys()]
    has_image_cols = any(h.startswith("image_") for h in headers)

    items: list[BatchItemInput] = []
    for raw_row in raw_rows:
        row = {k.strip().lower(): v.strip() for k, v in raw_row.items()}
        product_id = row.get("product_id") or row.get("id") or ""
        product_name = row.get("product_name") or row.get("title") or row.get("name") or ""
        if not product_id or not product_name:
            continue

        matched_images: list[str] = []
        if has_image_cols:
            for h in headers:
                if h.startswith("image_") and row.get(h):
                    matched_images.append(row[h])
        else:
            # prefix match
            matched_images = [img for img in available_images if Path(img).stem.startswith(product_id)]

        items.append(
            BatchItemInput(
                product_id=product_id,
                product_name=product_name,
                price=row.get("price", ""),
                category=row.get("category", "default") or "default",
                script_template=row.get("script_template", ""),
                image_names=matched_images if matched_images else None,
            )
        )

    # Rule 3 fallback: if every row got 0 matches and no explicit columns, pool all images
    if not has_image_cols and all((it.image_names is None or len(it.image_names) == 0) for it in items):
        for it in items:
            it.image_names = available_images.copy()

    return items


def create_zip_for_batch(batch_id: str, runs_dir: Path) -> Path:
    """Pack all final_with_voice.mp4 files from completed items into a ZIP."""
    conn = _get_conn()
    rows = conn.execute(
        "SELECT product_id, deliverables_dir FROM batch_items WHERE batch_id = ? AND status = 'completed' ORDER BY rowid",
        (batch_id,),
    ).fetchall()
    conn.close()

    zip_path = runs_dir.parent / "batch_downloads" / f"{batch_id}.zip"
    zip_path.parent.mkdir(parents=True, exist_ok=True)

    seen_ids: dict[str, int] = {}
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for row in rows:
            deliverables_dir = Path(row["deliverables_dir"]) if row["deliverables_dir"] else None
            if not deliverables_dir or not deliverables_dir.exists():
                continue
            video = deliverables_dir / "final_with_voice.mp4"
            if video.exists():
                pid = row["product_id"] or "unknown"
                count = seen_ids.get(pid, 0)
                seen_ids[pid] = count + 1
                arcname = f"{pid}_{count + 1}.mp4" if count > 0 else f"{pid}.mp4"
                zf.write(video, arcname)

    return zip_path
